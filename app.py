# Python Tutorial with "Mosh Hamedani"

# course = "Python for Beginners"
# print(course[0:5])
# print(len(course))  # to count number of character include space
# print(course.upper())
# print(course.lower())
# print(course.title())
# print(course.find("Beginners"))  # Return the index of "B"
# print('Python' in course)  # Return Boolean values, True or False
# print(course.replace("Beginners", "Absolute Beginners"))
#
# a = "This is the island of istanbul"
# # 3 is the maximum replacement that can be done in the string
# print(a.replace("is", "was", 3))  # Output --> Thwas was the wasland of istanbul


# If clause
# Score range 300-850, Bad credit is below 670
# credit_score = input('What is you "Credit Score"? ')
# print()
# credit_score = int(credit_score)
# price = 1000000
# if credit_score in range(670, 850):
#     down_payment = 0.1
#     down_payment *= price
#     print(f'You have Good Credit, so your down payment is only ${down_payment:.2f}')
# elif credit_score in range(300, 669):
#     down_payment = 0.2
#     down_payment *= price
#     print(f'You have Bad Credit, so you have to put down ${down_payment:.2f}')
# else:
#     print('Error! "Invalid Credit Score" \nPlease, type in your Credit Score between 300 - 850')


# ** Weight Converter **
# from ecommerce.converter import weight_converter
# weight = input('Weight: ')
# weight_converter(weight)

# **for loop**
# for item in 'Python':
#     print(item)
# print()
#
# for item in ['Jiab', 'Peter', 'Prateep']:
#     print(item)
# print()
#
# for item in range(10):
#     print(item)
#
# print()
# for item in range(5, 10):
#     print(item)

# # Calculate total price
# prices = [10, 20, 30]
# total = 0
# for price in prices:
#     total += (price)
# print(f'Total: {total}')
# print()
#
#
# # Nested loop
# for x in range(4):
#     for y in range(3):
#         print(f'({x},{y})')
#
# print()


# # Use for loop print "F" letter
# numbers = [2, 2, 2, 2, 8]
# for x_count in numbers:
#     count = ''
#     for i in range(x_count):
#         count += 'X'
#     print(count)
#
# print()

# # List
# names = ['Jiab', 'Peter', 'Prateep', 'Kul', 'Nattcha', 'ปิติ']
# print(names[1])
# print(names[-1])

# # finding the largest number in the list
# numbers = [3, 6, 8, 37, 5, 27]
# max_number = numbers[0]  # assume first number in the list is the largest number
# for number in numbers:
#     if number > max_number:
#         max_number = number
# print(max_number)


# # 2D List
# matrix = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
# matrix[0][1] = 20
# print(matrix[0][1])
# print(matrix)

# # Program to remove duplicate number in list
# numbers = [2, 4, 6, 8, 2, 4, 6, 10]
# unique = []  # Create new list if the number is not already in, add it to the list
# for number in numbers:
#     if number not in unique:
#         unique.append(number)
# print(unique)


# # Dictionary
# num_dictionary = {
#     '1': 'One',
#     '2': 'Two',
#     '3': 'Three',
#     '4': 'Four',
#     '5': 'Five',
#     '6': 'Six',
#     '7': 'Seven',
#     '8': 'Eight',
#     '9': 'Nine',
#     '0': 'Zero'
# }
#
# phone = input('Phone: ')
# converted_num = ''
# for number in phone:
#     converted_num += num_dictionary.get(number, '!') + ' '
# print(converted_num)

# # Emoji Function
# def emoji_converter(message):
#     words = message.split(' ')
#     emoji_dict = {
#         ':)': '😀',
#         ':(': '😔'
#     }
#     output = ''
#     for word in words:
#         output += emoji_dict.get(word, word) + ' '
#     return output
#
#
# message = input('> ')
# converted = emoji_converter(message)
# print(converted)


# ** Class **
# class Person:
#     def __init__(self, name):
#         self.name = name
#
#     def talk(self):
#         print(f'Hi, I am {self.name}')
#
#
# jiab = Person('Jiab Nattcha')
# jiab.talk()

# class Mammal:
#     def walk(self):
#         print('Walk')
#
#
# class Dog(Mammal):
#     def bark(self):
#         print('Bark')
#
#
# class Cat(Mammal):
#     pass
#
#
# dog1 = Dog()
# dog1.walk()

# from ecommerce.utils import find_max
# numbers = [3, 5, 8, 12, 2, 10, 6, 9]
# print(find_max(numbers))

# import random
#
#
# class Dice:
#     def roll(self):
#         dice1 = random.randint(1, 6)
#         dice2 = random.randint(1, 6)
#         return dice1, dice2
#
#
# dice = Dice()
# print(dice.roll())


# files and directory
# from pathlib import Path
# # Absolute path  --> start from root of hardisk
# # Relative path  --> start from current directory
# path = Path()
# for file in path.glob('*'):
#     print(file)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet.cell(1, 1)  # cell = sheet['a1']
    # print(cell.value)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
